from tkinter import *
from tkcalendar import DateEntry
import pandas as pd
import openpyxl


root = Tk()
root.geometry('610x570')

f1 = Frame(root, bg="burlywood1", height=1080, width=1920, pady=35)
f1.pack(fill=BOTH)

Label(f1, text="Select Date:", font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=0, column=0, padx=50, pady=5)
sel = StringVar()
cal = DateEntry(f1, locale='en_US', date_pattern='dd/mm/yyyy',
                selectmode='day', textvariable=sel, font="comicsansms 15", bg="grey85")
cal.grid(row=0, column=1, padx=20, pady=5)


def my_upd(*args):  # triggered when value of string varaible changes
    a = sel.get()  # read and display date


sel.trace('w', my_upd)


def green_colour_cell():

    from openpyxl.styles import PatternFill

    read = pd.read_excel("Trade_data.xlsx")
    wb = openpyxl.load_workbook("Trade_data.xlsx")

    ws = wb['Sheet1']
    fill_pattern = PatternFill(patternType='solid', fgColor='83F7A6')

    ls = []
    for index, val in enumerate(read['Result']):

        if val > 0:
            ls.append(f"J{index+2}")

    for i in ls:
        ws[f"{i}"].fill = fill_pattern
        wb.save("Trade_data.xlsx")


def red_colour_cell():

    from openpyxl.styles import PatternFill

    read2 = pd.read_excel("Trade_data.xlsx")
    wb2 = openpyxl.load_workbook("Trade_data.xlsx")

    ws2 = wb2['Sheet1']
    fill_pattern2 = PatternFill(patternType='solid', fgColor='FF8B7F')

    ls2 = []
    for index, val in enumerate(read2['Result']):

        if val < 0:
            ls2.append(f"J{index+2}")

    for i in ls2:
        ws2[f"{i}"].fill = fill_pattern2
        wb2.save("Trade_data.xlsx")


def submit_data():

    df = pd.DataFrame([[sel.get(), menu.get(), index.get(), strikeprice_value.get(), quantity_value.get(
    ), buy_time_value.get(), buy_value_value.get(), sell_time_value.get(), sell_value_value.get(), result_value.get()]])
    df.to_csv('Trade_csv_data.csv', mode='a', index=False, header=False)

    df = pd.DataFrame([[sel.get(), menu.get(), index.get(), strikeprice_value.get(), quantity_value.get(
    ), buy_time_value.get(), buy_value_value.get(), sell_time_value.get(), sell_value_value.get(), result_value.get()]])
    with pd.ExcelWriter("Trade_data.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name="Sheet1", header=None,
                    startrow=writer.sheets["Sheet1"].max_row, index=False)

    green_colour_cell()
    red_colour_cell()
    last_lab = Label(f1, text='Your Data has been submitted', bg="burlywood1")
    last_lab.grid(row=11, column=1, pady=5, padx=20)


menu = StringVar()
menu.set("Select Day")

# Create a dropdown Menu
Label(f1, text='Select Day:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=1, column=0, padx=20, pady=5)
drop = OptionMenu(f1, menu, "Monday", "Tuesday", "Wednesday",
                  "Thursday", "Friday", "Satuday")
drop.config(font="comicsansms 15", bg="grey85")
drop.grid(row=1, column=1, padx=20, pady=5)


index = StringVar()
index.set("Select Index:")

# Create a dropdown Menu
Label(f1, text='Select Index:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=2, column=0, padx=20, pady=5)
drop = OptionMenu(f1, index, "Bank Nifty", "Nifty")
drop.config(font="comicsansms 15", bg="grey85")
drop.grid(row=2, column=1, padx=20, pady=5)


Label(f1, text='Strike Price:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=3, column=0, padx=20, pady=5)
strikeprice_value = StringVar()
strikeprice_entry = Entry(f1, textvariable=strikeprice_value,
                          font="comicsansms 15", bg="grey85").grid(row=3, column=1, padx=20, pady=5)


Label(f1, text='Quantity:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=4, column=0, padx=20, pady=5)
quantity_value = StringVar()
quantity_entry = Entry(f1, textvariable=quantity_value, font="comicsansms 15",
                       bg="grey85").grid(row=4, column=1, padx=20, pady=5)


Label(f1, text='Buy Time:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=5, column=0, padx=20, pady=5)
buy_time_value = StringVar()
buy_time_entry = Entry(f1, textvariable=buy_time_value, font="comicsansms 15",
                       bg="grey85").grid(row=5, column=1, padx=20, pady=5)


Label(f1, text='Buy Value:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=6, column=0, padx=20, pady=5)
buy_value_value = StringVar()
buy_value_entry = Entry(f1, textvariable=buy_value_value, font="comicsansms 15",
                        bg="grey85").grid(row=6, column=1, padx=20, pady=5)


Label(f1, text='Sell Time:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=7, column=0, padx=20, pady=5)
sell_time_value = StringVar()
sell_time_entry = Entry(f1, textvariable=sell_time_value, font="comicsansms 15",
                        bg="grey85").grid(row=7, column=1, padx=20, pady=5)


Label(f1, text='Sell Value:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=8, column=0, padx=20, pady=5)
sell_value_value = StringVar()
sell_value_entry = Entry(f1, textvariable=sell_value_value, font="comicsansms 15",
                         bg="grey85").grid(row=8, column=1, padx=20, pady=5)


Label(f1, text='Result:', font="comicsansms 18 bold", bg="burlywood1",
      fg="midnight blue").grid(row=9, column=0, padx=20, pady=5)
result_value = StringVar()
result_entry = Entry(f1, textvariable=result_value, font="comicsansms 15",
                     bg="grey85").grid(row=9, column=1, padx=20, pady=5)


Button(f1, text='Submit', command=submit_data, font=("COPPERPLATE GOTHIC BOLD", 15,
       "bold"), bg="orange2", fg="grey19", relief=SUNKEN).grid(row=10, column=1, pady=5)

root.title('AnalyzingSoftware')
root.mainloop()
